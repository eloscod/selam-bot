import os
import logging
import json
import random
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
import telebot
from telebot import types
import time
from threading import Timer

# === Load Environment Variables ===
load_dotenv()
BOT_TOKEN = os.getenv('BOT_TOKEN')
ADMIN_ID = os.getenv('ADMIN_ID')

# === Logging Setup ===
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bot.log'),
        logging.StreamHandler()
    ]
)

# === Configurations ===
BASE_PATH = 'data/'
DATA_START_ROW = 5
DATA_END_ROW = 64
USER_MAPPING_FILE = 'user_mapping.json'
STUDENT_IDENTIFIERS_FILE = 'student_identifiers.json'
PROCESSED_CALLBACKS = set()
REGISTRATION_TIMEOUT = 300  # 5 minutes in seconds

# === Initialize Bot ===
bot = telebot.TeleBot(BOT_TOKEN)

# === Temporary Registration Storage ===
temp_registrations = {}

# === Localization Dictionary ===
MESSAGES = {
    'en': {
        'welcome': "🎓 *Welcome to Selam Islamic Elementary School Result Bot* 🎓\n--------------------------------\n📚 Official Bot for Selam Islamic Elementary School\n🌐 Serving Grades 1-6 with Real-Time Results\n👇 Click below to get started:\n\n- Check your individual results\n- View top 3 performing students\n\n📋 *Note:* Use /help for more information.",
        'not_authenticated': "🎓 *Selam Islamic Elementary School Result Bot* 🎓\n--------------------------------\nPlease register or log in to access your results:\n- Use `/register <grade_section> <student_no>` to get a PIN.\n- Use `/login <PIN>` to log in.\nContact admin for assistance.",
        'help': "🎓 *Selam Islamic Elementary School Result Bot Help* 🎓\n--------------------------------\n| Command       | Description                          |\n|---------------|--------------------------------------|\n| `/start`      | Start the bot                        |\n| `/help`       | Show this help message               |\n| `/register <grade_section> <student_no>` | Register to get a PIN |\n| `/login <PIN>` | Log in with your PIN to access results |\n| *Check My Results* | Select semester to view your results (after login) |\n| *View Top 3*  | Select section, semester, and view top 3 students |\n--------------------------------\nℹ️ *Notes:*\n- Register with `/register <grade_section> <student_no>` to get a PIN.\n- Log in with `/login <PIN>` to access your results.\n- Each Telegram account can only access one student's results.\n- Contact admin if you encounter issues.",
        'invalid_command': "❌ *Error:* Unrecognized command or input. Use /start, /register, /login, or /help.",
        'register_usage': "Usage: /register <grade_section> <student_no> (e.g., /register 1A 10)",
        'invalid_grade_section': "❌ *Error:* Invalid grade/section. Use: {sections}.",
        'invalid_student_no': "❌ *Error:* Invalid student number. Must be between 1 and 60.",
        'already_registered': "❌ *Error:* This Telegram account is already registered for Grade {grade_section}, Student No {student_no}.",
        'student_taken': "❌ *Error:* Student {student_no} in {grade_section} is already registered by another user.",
        'pin_error': "❌ *Error:* {error}. Contact admin.",
        'pin_sent': "✅ *Success:* Your PIN has been sent to your private chat. Check your direct messages from the bot and use /login <PIN> to proceed.",
        'pin_failed': "❌ *Error:* Failed to send PIN to your private chat. Ensure you have started a private conversation with the bot.",
        'register_success': "✅ *Registration Successful*\nYour PIN is: `{pin}`\nGrade: {grade_section}, Student No: {student_no}\nUse `/login {pin}` to access your results. Keep this PIN secure!",
        'login_usage': "Usage: /login <PIN> (e.g., /login 123456)",
        'invalid_pin': "❌ *Error:* Invalid PIN. Please check your PIN and try again.",
        'pin_not_owned': "❌ *Error:* This PIN is registered to another Telegram account.",
        'login_success': "✅ *Success:* Logged in for Grade {grade_section}, Student No {student_no}.\nYou can now use 'Check My Results' to view your results.",
        'not_logged_in': "❌ *Error:* You are not logged in. Use /register to get a PIN, then /login <PIN>.",
        'unauthorized_results': "❌ *Error:* You can only access your own results. Use /login <PIN>.",
        'select_grade_section': "📋 *Select Grade and Section* 🎓",
        'select_top3_section': "🏆 *Select Section for Top 3* 🎓",
        'selection_confirmed': "📋 *Selection Confirmed* 🎓\n✅ Grade {grade}, Section {section}\n--------------------------------\n🌟 Please select the semester:",
        'invalid_section': "❌ *Error:* Invalid section. Use: {sections}.",
        'file_not_found': "📁 *Error:* File for `{grade_section}` not found. Contact admin.",
        'sheet_not_found': "🗂️ *Error:* Sheet `{semester}` not found in {grade_section}.xlsx.",
        'invalid_excel': "⚠️ *Warning:* Invalid Excel structure for {grade_section} - {semester}. Please contact the admin.",
        'student_not_found': "❌ *Error:* Your student number `{student_no}` not found in {grade_section} - {semester}. Please contact the admin.",
        'no_averages': "⚠️ *Warning:* No averages found for {section} - {semester}.",
        'unexpected_error': "🚫 *Error:* An unexpected issue occurred. Details: {error}. Contact admin for support.",
        'result_header': "📄 *Your Result - {semester}* 🎓",
        'top3_header': "🏆 *Top 3 Students - {section}, {semester}* 🎓",
        'results_displayed': "✅ *Results Displayed Successfully!*",
        'check_my_results': "✅ Check My Results",
        'view_top3': "✅ View Top 3",
        'back_button': "⬅️ Back",
        'language_selection': "🌍 *Please select your preferred language:*",
        'registration_complete': "✅ Registration complete! You can now use the bot in English.",
        'registration_timeout': "⏳ Registration session expired. Please start again with /register.",
        'language_set': "✅ Language set to {language}."
    },
    'am': {
        'welcome': "🎓 *እንኳን ወደ ሰላም እስላማዊ አንደኛ ደረጃ ትምህርት ቤት ውጤት ቦት ተግባቢ እንኳን ደህና መጡ* 🎓\n--------------------------------\n📚 ለሰላም እስላማዊ አንደኛ ደረጃ ትምህርት ቤት ተግባቢ ቦት\n🌐 ከ1-6 ኛ ክፍል ውጤቶችን በእውነተኛ ጊዜ ያቀርባል\n👇 ለመጀመር ከታች ይጫኑ፡\n\n- የግል ውጤቶችዎን ይመልከቱ\n- ከፍተኛ 3 ተማሪዎችን ይመልከቱ\n\n📋 *ማሳሰቢያ:* ተጨማሪ መረጃ ለማግኘት /help ይጠቀሙ።",
        'not_authenticated': "🎓 *ሰላም እስላማዊ አንደኛ ደረጃ ትምህርት ቤት ውጤት ቦት* 🎓\n--------------------------------\nውጤቶችዎን ለመድረስ እባክዎ ይመዝገቡ ወይም ይግቡ፡\n- ፒን ለማግኘት `/register <grade_section> <student_no>` ይጠቀሙ።\n- ለመግባት `/login <PIN>` ይጠቀሙ።\nእርዳታ ለማግኘት አስተዳዳሪውን ያነጋግሩ።",
        'help': "🎓 *ሰላም እስላማዊ አንደኛ ደረጃ ትምህርት ቤት ውጤት ቦት እገዛ* 🎓\n--------------------------------\n| ትእዛዝ       | መግለጫ                             |\n|---------------|--------------------------------------|\n| `/start`      | ቦቱን ያስጀምሩ                    |\n| `/help`       | ይህን የእገዛ መልዕክት ያሳያል   |\n| `/register <grade_section> <student_no>` | ፒን ለማግኘት ይመዝገቡ |\n| `/login <PIN>` | ውጤቶችዎን ለመድረስ በፒን ይግቡ |\n| *ውጤቶቼን ይመልከቱ* | ሴሚስተር ይምረጡ እና ውጤቶችዎን ይመልከቱ (ከግቢያ በኋላ) |\n| *ከፍተኛ 3 ይመልከቱ*  | ክፍል፣ ሴሚስተር ይምረጡ እና ከፍተኛ 3 ተማሪዎችን ይመልከቱ |\n--------------------------------\nℹ️ *ማሳሰቢያ:*\n- ፒን ለማግኘት `/register <grade_section> <student_no>` ይጠቀሙ።\n- ውጤቶችዎን ለመድረስ `/login <PIN>` ይጠቀሙ።\n- እያንዳንዱ የቴሌግራም መለያ አንድ ተማሪ ውጤት ብቻ መድረስ ይችላል።\n- ችግር ካጋጠመዎት አስተዳዳሪውን ያነጋግሩ።",
        'invalid_command': "❌ *ስህተት:* ያልታወቀ ትእዛዝ ወይም ግብዓት። /start፣ /register፣ /login፣ /lang ወይም /help ይጠቀሙ።",
        'register_usage': "አጠቃቀም: /register <grade_section> <student_no> (ለምሳሌ፣ /register 1A 10)",
        'invalid_grade_section': "❌ *ስህተት:* የማይሰራ ክፍል/ክፍል። ይጠቀሙ: {sections}።",
        'invalid_student_no': "❌ *ስህተት:* የማይሰራ የተማሪ ቁጥር። ከ1 እስከ 60 መሆን አለበት።",
        'already_registered': "❌ *ስህተት:* ይህ የቴሌግራም መለያ ለ{grade_section} ክፍል፣ ተማሪ ቁጥር {student_no} ተመዝግቧል።",
        'student_taken': "❌ *ስህተት:* ተማሪ {student_no} በ{grade_section} ቀድሞ በሌላ ተጠቃሚ ተመዝግቧል።",
        'pin_error': "❌ *ስህተት:* {error}። አስተዳዳሪውን ያነጋግሩ።",
        'pin_sent': "✅ *ስኬት:* የእርስዎ ፒን ወደ ግል ውይይትዎ ተልኳል። ከቦቱ የተላኩ ቀጥታ መልዕክቶችን ይመልከቱ እና /login <PIN> ይጠቀሙ።",
        'pin_failed': "❌ *ስህተት:* ፒን ወደ ግል ውይይት መላክ አልተሳካም። ከቦቱ ጋር ግላዊ ውይይት መጀመርዎን ያረጋግጡ።",
        'register_success': "✅ *ምዝገባ ተሳክቷል*\nየእርስዎ ፒን: `{pin}`\nክፍል: {grade_section}፣ ተማሪ ቁጥር: {student_no}\nውጤቶችዎን ለመድረስ `/login {pin}` ይጠቀሙ። ይህን ፒን በደህና ያስቀምጡ!",
        'login_usage': "አጠቃቀም: /login <PIN> (ለምሳሌ፣ /login 123456)",
        'invalid_pin': "❌ *ስህተት:* የማይሰራ ፒን። እባክዎ ፒንዎን ያረጋግጡ እና እንደገና ይሞክሩ።",
        'pin_not_owned': "❌ *ስህተት:* ይህ ፒን ለሌላ የቴሌግራም መለያ ተመዝግቧል።",
        'login_success': "✅ *ስኬት:* ለ{grade_section} ክፍል፣ ተማሪ ቁጥር {student_no} ገብተዋል።\nአሁን ውጤቶችዎን ለመመልከት 'ውጤቶቼን ይመልከቱ' መጠቀም ይችላሉ።",
        'not_logged_in': "❌ *ስህተት:* አልገቡም። ፒን ለማግኘት /register ይጠቀሙ፣ ከዚያ /login <PIN>።",
        'unauthorized_results': "❌ *ስህተት:* የራስዎን ውጤቶች ብቻ መድረስ ይችላሉ። /login <PIN> ይጠቀሙ።",
        'select_grade_section': "📋 *ክፍል እና ክፍል ይምረጡ* 🎓",
        'select_top3_section': "🏆 *ከፍተኛ 3 ለመመልከት ክፍል ይምረጡ* 🎓",
        'selection_confirmed': "📋 *ምርጫ ተረጋግጧል* 🎓\n✅ ክፍል {grade}፣ ክፍል {section}\n--------------------------------\n🌟 እባክዎ ሴሚስተር ይምረጡ:",
        'invalid_section': "❌ *ስህተት:* የማይሰራ ክፍል። ይጠቀሙ: {sections}።",
        'file_not_found': "📁 *ስህተት:* ለ`{grade_section}` ፋይል አልተገኘም። አስተዳዳሪውን ያነጋግሩ።",
        'sheet_not_found': "🗂️ *ስህተት:* ሉህ `{semester}` በ{grade_section}.xlsx ውስጥ አልተገኘም።",
        'invalid_excel': "⚠️ *ማስጠንቀቂያ:* ለ{grade_section} - {semester} የማይሰራ የኤክሴል መዋቅር። አስተዳዳሪውን ያነጋግሩ።",
        'student_not_found': "❌ *ስህተት:* የእርስዎ ተማሪ ቁጥር `{student_no}` በ{grade_section} - {semester} ውስጥ አልተገኘም። አስተዳዳሪውን ያነጋግሩ።",
        'no_averages': "⚠️ *ማስጠንቀቂያ:* ለ{section} - {semester} ምንም አማካይ አልተገኘም።",
        'unexpected_error': "🚫 *ስህተት:* ያልተጠበቀ ችግር ተከስቷል። ዝርዝር: {error}። አስተዳዳሪውን ያነጋግሩ።",
        'result_header': "📄 *ውጤቶችዎ - {semester}* 🎓",
        'top3_header': "🏆 *ከፍተኛ 3 ተማሪዎች - {section}፣ {semester}* 🎓",
        'results_displayed': "✅ *ውጤቶች በተሳካ ሁኔታ ታይተዋል!*",
        'check_my_results': "✅ ውጤቶቼን ይመልከቱ",
        'view_top3': "✅ ከፍተኛ 3 ይመልከቱ",
        'back_button': "⬅️ ተመለስ",
        'language_selection': "🌍 *እባክዎ የሚፈልጉትን ቋንቋ ይምረጡ:*",
        'registration_complete': "✅ ምዝገባ ተጠናቅቋል! አሁን ቦቱን በአማርኛ መጠቀም ይችላሉ።",
        'registration_timeout': "⏳ የምዝገባ ሂደት ጊዜው አልፏል። እባክዎ እንደገና በ/register ይጀምሩ።",
        'language_set': "✅ ቋንቋ �ስለ {language} ተዘጋጅቷል።"
    }
}

# === Utility Functions ===
def is_number(val):
    try:
        return float(val) if val else False
    except:
        return False

def get_value(cell):
    return cell.value if cell and hasattr(cell, 'value') and cell.value is not None else 'N/A'

def validate_excel_structure(ws, semester):
    expected_cols = 18 if semester in ['S1', 'S2'] else 17
    if ws.max_column < expected_cols:
        logging.warning(f"Excel file for {ws.title} has fewer columns than expected ({ws.max_column} < {expected_cols})")
        return False
    return True

def get_loading_message(chat_id, message_id):
    dots = ["⏳", "⏳.", "⏳..", "⏳..."]
    for i in range(4):
        bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=dots[i])
        time.sleep(0.5)
    return bot.edit_message_text(chat_id=chat_id, message_id=message_id, text="⏳ Processing...")

def generate_unique_pin(identifiers):
    max_attempts = 100
    for _ in range(max_attempts):
        pin = f"{random.randint(0, 999999):06d}"  # Generate 6-digit PIN
        if pin not in identifiers:
            return pin
    raise ValueError("Could not generate unique PIN after maximum attempts.")

def get_user_language(user_id):
    user_mapping = load_user_mapping()
    return user_mapping.get(user_id, {}).get('language', 'en')

def cleanup_temp_registration(user_id):
    if user_id in temp_registrations:
        del temp_registrations[user_id]
        logging.info(f"Cleaned up temp registration for user {user_id}")

def schedule_registration_cleanup(user_id):
    timer = Timer(REGISTRATION_TIMEOUT, cleanup_temp_registration, args=[user_id])
    timer.start()
    return timer

# === Admin Notification Functions ===
def notify_admin(message_text):
    if ADMIN_ID:
        try:
            bot.send_message(ADMIN_ID, message_text, parse_mode="Markdown")
            logging.info(f"Admin notified: {message_text}")
        except telebot.apihelper.ApiException as e:
            logging.error(f"Failed to notify admin: {str(e)}")

def notify_admin_on_restart():
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S EAT')
    notify_admin(f"🔔 *Bot Restarted* at {timestamp}")

def notify_admin_on_result_view(user_id, username, grade_section, semester, student_no, result_text):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S EAT')
    username_str = f"@{username}" if username else "No username"
    notify_admin(
        f"📊 *Student Result Viewed*\n"
        f"--------------------------------\n"
        f"👤 *User ID:* {user_id}\n"
        f"👤 *Username:* {username_str}\n"
        f"📋 *Grade/Section:* {grade_section}\n"
        f"📅 *Semester:* {semester}\n"
        f"🔢 *Student No:* {student_no}\n"
        f"🕒 *Time:* {timestamp}\n"
        f"--------------------------------\n"
        f"{result_text}"
    )

def notify_admin_on_registration(user_id, username, grade_section, student_no, pin):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S EAT')
    username_str = f"@{username}" if username else "No username"
    notify_admin(
        f"📋 *New Student Registration*\n"
        f"--------------------------------\n"
        f"👤 *User ID:* {user_id}\n"
        f"👤 *Username:* {username_str}\n"
        f"📋 *Grade/Section:* {grade_section}\n"
        f"🔢 *Student No:* {student_no}\n"
        f"🔑 *PIN:* {pin}\n"
        f"🕒 *Time:* {timestamp}"
    )

# === Load and Save Mappings ===
def load_user_mapping():
    try:
        with open(USER_MAPPING_FILE, 'r') as f:
            return json.load(f).get('users', {})
    except (FileNotFoundError, json.JSONDecodeError):
        logging.error(f"Error loading {USER_MAPPING_FILE}. Returning empty mapping.")
        return {}

def load_student_identifiers():
    try:
        with open(STUDENT_IDENTIFIERS_FILE, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        logging.error(f"Error loading {STUDENT_IDENTIFIERS_FILE}. Returning empty mapping.")
        return {}

def save_user_mapping(user_mapping):
    with open(USER_MAPPING_FILE, 'w') as f:
        json.dump({'users': user_mapping}, f, indent=2)

def save_student_identifiers(identifiers):
    with open(STUDENT_IDENTIFIERS_FILE, 'w') as f:
        json.dump(identifiers, f, indent=2)

# === Registration with Language Selection ===
@bot.message_handler(commands=['register'])
def register_user(message):
    user_id = str(message.from_user.id)
    lang = get_user_language(user_id)
    args = message.text.split()
    
    if len(args) != 3:
        bot.reply_to(message, MESSAGES[lang]['register_usage'], parse_mode="Markdown")
        return

    grade_section, student_no = args[1], args[2]
    username = message.from_user.username
    user_mapping = load_user_mapping()
    student_identifiers = load_student_identifiers()

    # Validate grade_section
    valid_sections = ['1A', '1B', '1C', '2A', '2B', '2C', '3A', '3B', '4A', '4B', '5A', '5B', '6A', '6B']
    if grade_section not in valid_sections:
        bot.reply_to(message, MESSAGES[lang]['invalid_grade_section'].format(sections=', '.join(valid_sections)), parse_mode="Markdown")
        return

    # Validate student_no
    if not student_no.isdigit() or not (1 <= int(student_no) <= 60):
        bot.reply_to(message, MESSAGES[lang]['invalid_student_no'], parse_mode="Markdown")
        return

    # Check if user is already registered
    if user_id in user_mapping:
        bot.reply_to(message, MESSAGES[lang]['already_registered'].format(
            grade_section=user_mapping[user_id]['grade_section'],
            student_no=user_mapping[user_id]['student_no']
        ), parse_mode="Markdown")
        return

    # Check if student is already registered by another user
    for pin, data in student_identifiers.items():
        if data['grade_section'] == grade_section and data['student_no'] == student_no:
            bot.reply_to(message, MESSAGES[lang]['student_taken'].format(
                student_no=student_no,
                grade_section=grade_section
            ), parse_mode="Markdown")
            return

    # Store temporary registration data
    temp_registrations[user_id] = {
        'grade_section': grade_section,
        'student_no': student_no,
        'username': username,
        'message_id': message.message_id,
        'timer': schedule_registration_cleanup(user_id)
    }

    # Prompt for language selection
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("English 🇬🇧", callback_data=f'reg_lang_en_{grade_section}_{student_no}'),
        types.InlineKeyboardButton("Amharic 🇪🇹", callback_data=f'reg_lang_am_{grade_section}_{student_no}')
    )
    
    bot.reply_to(message, 
        MESSAGES[lang]['language_selection'],
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data.startswith('reg_lang_'))
def handle_registration_language(call):
    user_id = str(call.from_user.id)
    _, _, lang, grade_section, student_no = call.data.split('_', 4)
    
    if user_id not in temp_registrations:
        bot.answer_callback_query(call.id, MESSAGES['en']['registration_timeout'])
        return

    registration_data = temp_registrations[user_id]
    registration_data['timer'].cancel()  # Cancel the cleanup timer
    
    username = registration_data['username']
    user_mapping = load_user_mapping()
    student_identifiers = load_student_identifiers()

    # Generate unique 6-digit PIN
    try:
        pin = generate_unique_pin(student_identifiers)
    except ValueError as e:
        bot.send_message(call.message.chat.id, MESSAGES[lang]['pin_error'].format(error=str(e)), parse_mode="Markdown")
        return

    # Store student data with Telegram ID
    student_identifiers[pin] = {
        'grade_section': grade_section,
        'student_no': student_no,
        'telegram_id': user_id
    }
    save_student_identifiers(student_identifiers)

    # Store user mapping with language preference
    user_mapping[user_id] = {
        'grade_section': grade_section,
        'student_no': student_no,
        'pin': pin,
        'language': lang
    }
    save_user_mapping(user_mapping)

    # Clean up temporary data
    del temp_registrations[user_id]

    # Send confirmation
    try:
        bot.send_message(user_id, MESSAGES[lang]['register_success'].format(
            pin=pin,
            grade_section=grade_section,
            student_no=student_no
        ), parse_mode="Markdown")
        
        bot.edit_message_text(
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            text=MESSAGES[lang]['registration_complete'],
            reply_markup=None
        )
        
        notify_admin_on_registration(user_id, username, grade_section, student_no, pin)
        show_welcome_message(call.message)  # Show welcome message in selected language
        
    except telebot.apihelper.ApiException as e:
        bot.send_message(call.message.chat.id, MESSAGES[lang]['pin_failed'], parse_mode="Markdown")
        logging.error(f"Failed to send PIN to {user_id}: {str(e)}")

# === /login Command Handler ===
@bot.message_handler(commands=['login'])
def login_user(message):
    user_id = str(message.from_user.id)
    lang = get_user_language(user_id)
    args = message.text.split()
    if len(args) != 2:
        bot.reply_to(message, MESSAGES[lang]['login_usage'], parse_mode="Markdown")
        return

    pin = args[1]
    username = message.from_user.username
    user_mapping = load_user_mapping()
    student_identifiers = load_student_identifiers()

    if pin not in student_identifiers:
        bot.reply_to(message, MESSAGES[lang]['invalid_pin'], parse_mode="Markdown")
        return

    student_data = student_identifiers[pin]
    if student_data['telegram_id'] != user_id:
        bot.reply_to(message, MESSAGES[lang]['pin_not_owned'], parse_mode="Markdown")
        return

    # Update user mapping
    user_mapping[user_id] = {
        'grade_section': student_data['grade_section'],
        'student_no': student_data['student_no'],
        'pin': pin,
        'language': lang  # Preserve existing language
    }
    save_user_mapping(user_mapping)
    bot.reply_to(message, MESSAGES[lang]['login_success'].format(
        grade_section=student_data['grade_section'],
        student_no=student_data['student_no']
    ), parse_mode="Markdown")
    show_welcome_message(message)  # Show welcome message after successful login

# === /start Command Handler ===
@bot.message_handler(commands=['start'])
def send_welcome(message):
    user_id = str(message.from_user.id)
    lang = get_user_language(user_id)
    user_mapping = load_user_mapping()

    if user_id in user_mapping:
        show_welcome_message(message)
    else:
        bot.reply_to(message, MESSAGES[lang]['not_authenticated'], parse_mode="Markdown")

# === Show Welcome Message ===
def show_welcome_message(message):
    user_id = str(message.from_user.id)
    lang = get_user_language(user_id)
    max_retries = 3
    for attempt in range(max_retries):
        try:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton(MESSAGES[lang]['check_my_results'], callback_data='results'))
            markup.add(types.InlineKeyboardButton(MESSAGES[lang]['view_top3'], callback_data='top3'))
            bot.reply_to(message, MESSAGES[lang]['welcome'], parse_mode="Markdown", reply_markup=markup)
            break
        except telebot.apihelper.ApiException as e:
            logging.error(f"Attempt {attempt + 1} failed: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
            else:
                bot.reply_to(message, MESSAGES[lang]['unexpected_error'].format(error="Unable to connect to Telegram"), parse_mode="Markdown")

# === /help Command Handler ===
@bot.message_handler(commands=['help'])
def send_help(message):
    user_id = str(message.from_user.id)
    lang = get_user_language(user_id)
    bot.reply_to(message, MESSAGES[lang]['help'], parse_mode="Markdown")

# === /lang Command Handler ===
@bot.message_handler(commands=['lang'])
def set_language(message):
    user_id = str(message.from_user.id)
    current_lang = get_user_language(user_id)
    args = message.text.split()
    
    if len(args) != 2:
        bot.reply_to(message, MESSAGES[current_lang]['language_selection'], parse_mode="Markdown")
        return

    lang = args[1].lower()
    if lang not in ['am', 'en']:
        bot.reply_to(message, MESSAGES[current_lang]['invalid_lang'], parse_mode="Markdown")
        return

    user_mapping = load_user_mapping()
    if user_id in user_mapping:
        user_mapping[user_id]['language'] = lang
    else:
        user_mapping[user_id] = {'language': lang}
    
    save_user_mapping(user_mapping)
    
    lang_name = "Amharic" if lang == "am" else "English"
    bot.reply_to(message, MESSAGES[lang]['language_set'].format(language=lang_name), parse_mode="Markdown")

# === Catch Unexpected Input ===
@bot.message_handler(func=lambda message: True)
def handle_unexpected_input(message):
    user_id = str(message.from_user.id)
    lang = get_user_language(user_id)
    bot.reply_to(message, MESSAGES[lang]['invalid_command'], parse_mode="Markdown")

# === Helper Functions for Markups ===
def get_grade_section_markup(is_top3=False, lang='en'):
    markup = types.InlineKeyboardMarkup(row_width=4)
    sections = ['1A', '1B', '1C', '2A', '2B', '2C', '3A', '3B', '4A', '4B', '5A', '5B', '6A', '6B']
    for i in range(0, len(sections), 4):
        row = sections[i:i + 4]
        buttons = [types.InlineKeyboardButton(f"✅ {sec}", callback_data=f'grade_{"top3" if is_top3 else ""}{sec}') for sec in row]
        if i == 0:
            buttons.append(types.InlineKeyboardButton(MESSAGES[lang]['back_button'], callback_data=f'grade_{"top3" if is_top3 else ""}{row[0]}_back'))
        markup.row(*buttons[:4])
    return markup

def get_semester_markup(grade_section, is_top3=False, lang='en'):
    markup = types.InlineKeyboardMarkup(row_width=4)
    markup.add(
        types.InlineKeyboardButton("✅ S1", callback_data=f'{"semester_" if not is_top3 else "top3_"}{grade_section}_S1'),
        types.InlineKeyboardButton("✅ S2", callback_data=f'{"semester_" if not is_top3 else "top3_"}{grade_section}_S2'),
        types.InlineKeyboardButton("✅ Ave", callback_data=f'{"semester_" if not is_top3 else "top3_"}{grade_section}_Ave'),
        types.InlineKeyboardButton(MESSAGES[lang]['back_button'], callback_data=f'{"semester_" if not is_top3 else "top3_"}{grade_section}_back')
    )
    return markup

def prompt_grade_section(message, is_top3=False):
    user_id = str(message.from_user.id)
    lang = get_user_language(user_id)
    bot.reply_to(message,
        MESSAGES[lang]['select_grade_section'] if not is_top3 else MESSAGES[lang]['select_top3_section'],
        reply_markup=get_grade_section_markup(is_top3, lang),
        parse_mode="Markdown"
    )

def prompt_semester(grade_section, lang='en'):
    return get_semester_markup(grade_section, is_top3=False, lang=lang)

def prompt_top3_semester(grade_section, lang='en'):
    return get_semester_markup(grade_section, is_top3=True, lang=lang)

# === Handle Inline Keyboard Callbacks ===
@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    global PROCESSED_CALLBACKS
    if call.id in PROCESSED_CALLBACKS:
        bot.answer_callback_query(call.id, "Request already processed.")
        return
    PROCESSED_CALLBACKS.add(call.id)

    user_id = str(call.from_user.id)
    lang = get_user_language(user_id)
    user_mapping = load_user_mapping()

    if call.data == 'results':
        bot.answer_callback_query(call.id)
        if user_id not in user_mapping:
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=MESSAGES[lang]['not_logged_in'],
                parse_mode="Markdown"
            )
            return
        grade_section = user_mapping[user_id]['grade_section']
        bot.edit_message_text(
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            text=MESSAGES[lang]['selection_confirmed'].format(grade=grade_section[0], section=grade_section[1]),
            reply_markup=prompt_semester(grade_section, lang),
            parse_mode="Markdown"
        )
    elif call.data == 'top3':
        bot.answer_callback_query(call.id)
        prompt_grade_section(call.message, is_top3=True)
    elif call.data.startswith('grade_'):
        bot.answer_callback_query(call.id)
        grade_section = call.data.replace('grade_', '').replace('top3', '')
        if call.data.endswith('_back'):
            prompt_grade_section(call.message, is_top3=call.data.startswith('grade_top3'))
        else:
            next_step = prompt_semester if not call.data.startswith('grade_top3') else prompt_top3_semester
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=MESSAGES[lang]['selection_confirmed'].format(grade=grade_section[0], section=grade_section[1]),
                reply_markup=next_step(grade_section, lang),
                parse_mode="Markdown"
            )
    elif call.data.startswith(('semester_', 'top3_')) and not call.data.endswith('_back'):
        bot.answer_callback_query(call.id)
        if call.data.startswith('semester_'):
            grade_section, semester = call.data.replace('semester_', '').split('_')
            if user_id not in user_mapping or user_mapping[user_id]['grade_section'] != grade_section:
                bot.edit_message_text(
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    text=MESSAGES[lang]['unauthorized_results'],
                    parse_mode="Markdown"
                )
                return
            student_no = user_mapping[user_id]['student_no']
            process_results(call.message, grade_section, semester, student_no, user_id, call.from_user.username, lang)
        else:
            section, semester = call.data.replace('top3_', '').split('_')
            process_top3(call.message, section, semester, lang)
    elif call.data.endswith('_back'):
        bot.answer_callback_query(call.id)
        if call.data.startswith('semester_'):
            grade_section = call.data.replace('semester__back', '').split('_')[0]
            if user_id not in user_mapping or user_mapping[user_id]['grade_section'] != grade_section:
                bot.edit_message_text(
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    text=MESSAGES[lang]['unauthorized_results'],
                    parse_mode="Markdown"
                )
                return
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=MESSAGES[lang]['selection_confirmed'].format(grade=grade_section[0], section=grade_section[1]),
                reply_markup=prompt_semester(grade_section, lang),
                parse_mode="Markdown"
            )
        elif call.data.startswith('top3_'):
            prompt_grade_section(call.message, is_top3=True)

# === Process Results ===
def process_results(message, grade_section, semester, student_no, user_id, username, lang):
    loading_msg = bot.reply_to(message, "⏳ Processing...")
    try:
        grade = int(grade_section[0])
        section = grade_section[1]
        valid_sections = ['A', 'B', 'C'] if grade in [1, 2] else ['A', 'B']
        if section not in valid_sections:
            bot.reply_to(message, MESSAGES[lang]['invalid_section'].format(sections=', '.join(valid_sections)), parse_mode="Markdown")
            bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
            return

        file_path = f"{BASE_PATH}{grade_section}.xlsx"
        wb = load_workbook(file_path, data_only=True)
        ws = wb[semester]

        if not validate_excel_structure(ws, semester):
            bot.reply_to(message, MESSAGES[lang]['invalid_excel'].format(grade_section=grade_section, semester=semester), parse_mode="Markdown")
            bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
            return

        name_index = 3 if semester in ['S1', 'S2'] else 2
        max_col = 18 if semester in ['S1', 'S2'] else 17

        subjects = {
            'en': ["Amharic", "English", "Arabic", "Maths", "E.S", "Moral Edu", "Art", "HPE"],
            'am': ["አማርኛ", "እንግሊዝኛ", "አረብኛ", "ሒሳብ", "ኢ.ኤስ", "ሥነ ምግባር ትምህርት", "ሥነ ጥበብ", "ኤች.ፒ.ኢ"]
        }

        for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=DATA_END_ROW, max_col=max_col):
            row_no = str(get_value(row[1])).strip()
            if row_no == student_no:
                result_text = (
                    f"{MESSAGES[lang]['result_header'].format(semester=semester)}\n"
                    f"--------------------------------\n"
                    f"👤 *{'Student No' if lang == 'en' else 'የተማሪ ቁጥር'}:* {get_value(row[1])}\n"
                    f"👤 *{'Name' if lang == 'en' else 'ስም'}:* {get_value(row[name_index])}\n"
                    f"🔢 *{'Sex' if lang == 'en' else 'ፆታ'}:* {get_value(row[name_index + 1])}\n"
                    f"🎂 *{'Age' if lang == 'en' else 'ዕድሜ'}:* {get_value(row[name_index + 2])}\n"
                    f"📚 *{'Subjects' if lang == 'en' else 'ትምህርቶች'}:*\n"
                    f" - {subjects[lang][0]}: {get_value(row[name_index + 3])}\n"
                    f" - {subjects[lang][1]}: {get_value(row[name_index + 4])}\n"
                    f" - {subjects[lang][2]}: {get_value(row[name_index + 5])}\n"
                    f" - {subjects[lang][3]}: {get_value(row[name_index + 6])}\n"
                    f" - {subjects[lang][4]}: {get_value(row[name_index + 7])}\n"
                    f" - {subjects[lang][5]}: {get_value(row[name_index + 8])}\n"
                    f" - {subjects[lang][6]}: {get_value(row[name_index + 9])}\n"
                    f" - {subjects[lang][7]}: {get_value(row[name_index + 10])}\n"
                    f"💡 *{'Conduct' if lang == 'en' else 'ባህሪ'}:* {get_value(row[14]) if semester in ['S1', 'S2'] else 'N/A'}\n"
                    f"🧮 *{'Sum' if lang == 'en' else 'ድምር'}:* {get_value(row[15]) if semester in ['S1', 'S2'] else get_value(row[13])}\n"
                    f"📊 *{'Average' if lang == 'en' else 'አማካይ'}:* {get_value(row[16]) if semester in ['S1', 'S2'] else get_value(row[14])}\n"
                    f"🏅 *{'Rank' if lang == 'en' else 'ደረጃ'}:* {get_value(row[17]) if semester in ['S1', 'S2'] else get_value(row[15])}\n"
                    f"📝 *{'Remark' if lang == 'en' else 'አስተያየት'}:* {get_value(row[16]) if semester == 'Ave' else 'N/A'}\n"
                    f"--------------------------------\n"
                    f"{MESSAGES[lang]['results_displayed']}"
                )
                bot.reply_to(message, result_text, parse_mode="Markdown")
                notify_admin_on_result_view(user_id, username, grade_section, semester, student_no, result_text)
                time.sleep(1)
                bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
                return

        bot.reply_to(message, MESSAGES[lang]['student_not_found'].format(
            student_no=student_no,
            grade_section=grade_section,
            semester=semester
        ), parse_mode="Markdown")
        bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)

    except FileNotFoundError:
        bot.reply_to(message, MESSAGES[lang]['file_not_found'].format(grade_section=grade_section), parse_mode="Markdown")
        bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
    except KeyError:
        bot.reply_to(message, MESSAGES[lang]['sheet_not_found'].format(semester=semester, grade_section=grade_section), parse_mode="Markdown")
        bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
    except Exception as e:
        logging.exception("Unexpected error in /results")
        bot.reply_to(message, MESSAGES[lang]['unexpected_error'].format(error=str(e)), parse_mode="Markdown")
        bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)

# === Process Top 3 ===
def process_top3(message, section, semester, lang):
    loading_msg = bot.reply_to(message, "⏳ Processing...")
    try:
        grade = int(section[0])
        sec_code = section[1]
        valid_sections = ['A', 'B', 'C'] if grade in [1, 2] else ['A', 'B']
        if sec_code not in valid_sections:
            bot.reply_to(message, MESSAGES[lang]['invalid_section'].format(sections=', '.join(valid_sections)), parse_mode="Markdown")
            bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
            return
        file_path = f"{BASE_PATH}{section}.xlsx"
        wb = load_workbook(file_path, data_only=True)
        ws = wb[semester]

        if not validate_excel_structure(ws, semester):
            bot.reply_to(message, MESSAGES[lang]['invalid_excel'].format(grade_section=section, semester=semester), parse_mode="Markdown")
            bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
            return

        students = []
        for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=DATA_END_ROW, max_col=18):
            no, name, avg = row[1].value, row[3].value, row[16].value
            if no and avg and is_number(avg):
                students.append({'no': no, 'name': name, 'average': float(avg)})
        if not students:
            bot.reply_to(message, MESSAGES[lang]['no_averages'].format(section=section, semester=semester), parse_mode="Markdown")
            bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
            return
        top3 = sorted(students, key=lambda x: x['average'], reverse=True)[:3]
        response = f"{MESSAGES[lang]['top3_header'].format(section=section, semester=semester)}\n" + "\n".join([
            f"--------------------------------\n"
            f"{i+1}. 👤 *{'Name' if lang == 'en' else 'ስም'}:* {s['name']} ({'No' if lang == 'en' else 'ቁጥር'}: {s['no']}, 📊 *{'Avg' if lang == 'en' else 'አማካይ'}:* {s['average']:.1f})"
            for i, s in enumerate(top3)
        ]) + f"\n--------------------------------\n{MESSAGES[lang]['results_displayed']}"
        bot.reply_to(message, response, parse_mode="Markdown")
        time.sleep(1)
        bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
    except FileNotFoundError:
        bot.reply_to(message, MESSAGES[lang]['file_not_found'].format(grade_section=section), parse_mode="Markdown")
        bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
    except KeyError:
        bot.reply_to(message, MESSAGES[lang]['sheet_not_found'].format(semester=semester, grade_section=section), parse_mode="Markdown")
        bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)
    except Exception as e:
        logging.exception("Error in /top3")
        bot.reply_to(message, MESSAGES[lang]['unexpected_error'].format(error=str(e)), parse_mode="Markdown")
        bot.delete_message(chat_id=loading_msg.chat.id, message_id=loading_msg.message_id)

# === Run Bot ===
if __name__ == '__main__':
    logging.info("📡 Bot is running...")
    notify_admin_on_restart()
    bot.polling()